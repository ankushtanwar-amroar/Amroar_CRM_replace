"""
Excel Object Import Service
Handles parsing, validation, and creation of Custom Objects from Excel files.
"""
import io
import re
import uuid
from datetime import datetime, timezone
from typing import Dict, List, Any, Optional, Tuple
from openpyxl import load_workbook, Workbook
from motor.motor_asyncio import AsyncIOMotorDatabase
import logging

logger = logging.getLogger(__name__)

# Valid data types for fields
VALID_DATA_TYPES = {
    'text': {'mongo_type': 'string', 'default_component': 'text_input'},
    'number': {'mongo_type': 'number', 'default_component': 'number_input'},
    'date': {'mongo_type': 'date', 'default_component': 'date_picker'},
    'datetime': {'mongo_type': 'datetime', 'default_component': 'datetime_picker'},
    'boolean': {'mongo_type': 'boolean', 'default_component': 'checkbox'},
    'picklist': {'mongo_type': 'string', 'default_component': 'select'},
    'multipicklist': {'mongo_type': 'array', 'default_component': 'multi_select'},
    'lookup': {'mongo_type': 'string', 'default_component': 'lookup'},
    'textarea': {'mongo_type': 'string', 'default_component': 'textarea'},
    'email': {'mongo_type': 'string', 'default_component': 'email_input'},
    'phone': {'mongo_type': 'string', 'default_component': 'phone_input'},
    'url': {'mongo_type': 'string', 'default_component': 'url_input'},
    'currency': {'mongo_type': 'number', 'default_component': 'currency_input'},
    'percent': {'mongo_type': 'number', 'default_component': 'percent_input'},
}

# System fields that are auto-created
SYSTEM_FIELDS = {
    'created_at': {
        'label': 'Created Date',
        'type': 'datetime',
        'required': False,
        'system': True,
        'editable': False
    },
    'created_by': {
        'label': 'Created By',
        'type': 'lookup',
        'required': False,
        'system': True,
        'editable': False,
        'lookup_object': 'user'
    },
    'updated_at': {
        'label': 'Last Modified Date',
        'type': 'datetime',
        'required': False,
        'system': True,
        'editable': False
    },
    'updated_by': {
        'label': 'Last Modified By',
        'type': 'lookup',
        'required': False,
        'system': True,
        'editable': False,
        'lookup_object': 'user'
    },
    'owner_id': {
        'label': 'Owner',
        'type': 'lookup',
        'required': False,
        'system': True,
        'editable': True,
        'lookup_object': 'user'
    }
}


class ExcelObjectService:
    def __init__(self, db: AsyncIOMotorDatabase):
        self.db = db
        self.objects_collection = db.tenant_objects
        
    def generate_api_name(self, label: str) -> str:
        """Generate API name from label"""
        # Convert to lowercase, replace spaces with underscores
        api_name = label.lower().strip()
        api_name = re.sub(r'[^a-z0-9\s]', '', api_name)
        api_name = re.sub(r'\s+', '_', api_name)
        # Ensure it doesn't start with a number
        if api_name and api_name[0].isdigit():
            api_name = 'field_' + api_name
        return api_name
    
    def parse_excel(self, file_content: bytes) -> Tuple[Dict[str, Any], List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        Parse Excel file and extract object and field definitions.
        Returns: (object_data, fields_data, errors)
        """
        errors = []
        object_data = {}
        fields_data = []
        
        try:
            workbook = load_workbook(filename=io.BytesIO(file_content), data_only=True)
        except Exception as e:
            errors.append({
                'type': 'file_error',
                'message': f'Invalid Excel file format: {str(e)}',
                'sheet': None,
                'row': None,
                'column': None
            })
            return object_data, fields_data, errors
        
        # Check required sheets
        sheet_names = [s.lower() for s in workbook.sheetnames]
        if 'object' not in sheet_names:
            errors.append({
                'type': 'sheet_missing',
                'message': 'Required sheet "Object" not found',
                'sheet': 'Object',
                'row': None,
                'column': None
            })
        if 'fields' not in sheet_names:
            errors.append({
                'type': 'sheet_missing',
                'message': 'Required sheet "Fields" not found',
                'sheet': 'Fields',
                'row': None,
                'column': None
            })
        
        if errors:
            return object_data, fields_data, errors
        
        # Parse Object sheet
        object_sheet = None
        fields_sheet = None
        for sheet_name in workbook.sheetnames:
            if sheet_name.lower() == 'object':
                object_sheet = workbook[sheet_name]
            elif sheet_name.lower() == 'fields':
                fields_sheet = workbook[sheet_name]
        
        # Parse Object sheet
        object_data, obj_errors = self._parse_object_sheet(object_sheet)
        errors.extend(obj_errors)
        
        # Parse Fields sheet
        fields_data, field_errors = self._parse_fields_sheet(fields_sheet)
        errors.extend(field_errors)
        
        return object_data, fields_data, errors
    
    def _parse_object_sheet(self, sheet) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
        """Parse the Object sheet"""
        errors = []
        object_data = {}
        
        # Get headers from first row
        headers = {}
        for col_idx, cell in enumerate(sheet[1], 1):
            if cell.value:
                headers[cell.value.lower().strip().replace(' ', '_')] = col_idx
        
        # Required columns
        required_columns = ['object_label']
        for col in required_columns:
            if col not in headers:
                errors.append({
                    'type': 'column_missing',
                    'message': f'Required column "{col.replace("_", " ").title()}" not found',
                    'sheet': 'Object',
                    'row': 1,
                    'column': None
                })
        
        if errors:
            return object_data, errors
        
        # Check for data rows (should be exactly 1)
        data_rows = list(sheet.iter_rows(min_row=2, max_row=sheet.max_row))
        non_empty_rows = [row for row in data_rows if any(cell.value for cell in row)]
        
        if len(non_empty_rows) == 0:
            errors.append({
                'type': 'no_data',
                'message': 'No object data found in Object sheet',
                'sheet': 'Object',
                'row': 2,
                'column': None
            })
            return object_data, errors
        
        if len(non_empty_rows) > 1:
            errors.append({
                'type': 'too_many_rows',
                'message': 'Only one object can be defined per Excel file',
                'sheet': 'Object',
                'row': 3,
                'column': None
            })
        
        # Parse first data row
        row = non_empty_rows[0]
        row_data = {}
        for header_name, col_idx in headers.items():
            cell_value = row[col_idx - 1].value
            row_data[header_name] = cell_value.strip() if isinstance(cell_value, str) else cell_value
        
        # Validate required fields
        if not row_data.get('object_label'):
            errors.append({
                'type': 'required_field',
                'message': 'Object Label is required',
                'sheet': 'Object',
                'row': 2,
                'column': headers.get('object_label')
            })
        
        # Build object data
        object_label = row_data.get('object_label', '')
        plural_label = row_data.get('plural_label') or (object_label + 's' if object_label else '')
        api_name = row_data.get('api_name') or self.generate_api_name(object_label)
        
        object_data = {
            'object_label': object_label,
            'object_plural': plural_label,
            'object_name': api_name.lower(),
            'description': row_data.get('description', ''),
            'icon': row_data.get('icon', 'file')
        }
        
        return object_data, errors
    
    def _parse_fields_sheet(self, sheet) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """Parse the Fields sheet"""
        errors = []
        fields_data = []
        
        # Get headers from first row
        headers = {}
        for col_idx, cell in enumerate(sheet[1], 1):
            if cell.value:
                headers[cell.value.lower().strip().replace(' ', '_')] = col_idx
        
        # Required columns
        required_columns = ['field_label', 'data_type']
        for col in required_columns:
            if col not in headers:
                errors.append({
                    'type': 'column_missing',
                    'message': f'Required column "{col.replace("_", " ").title()}" not found',
                    'sheet': 'Fields',
                    'row': 1,
                    'column': None
                })
        
        if errors:
            return fields_data, errors
        
        # Parse data rows
        seen_api_names = set()
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), 2):
            # Skip empty rows
            if not any(cell.value for cell in row):
                continue
            
            row_data = {}
            for header_name, col_idx in headers.items():
                cell_value = row[col_idx - 1].value
                row_data[header_name] = cell_value.strip() if isinstance(cell_value, str) else cell_value
            
            field_errors = []
            
            # Validate required fields
            field_label = row_data.get('field_label', '')
            if not field_label:
                field_errors.append({
                    'type': 'required_field',
                    'message': 'Field Label is required',
                    'sheet': 'Fields',
                    'row': row_idx,
                    'column': headers.get('field_label')
                })
            
            data_type = str(row_data.get('data_type', '')).lower().strip()
            if not data_type:
                field_errors.append({
                    'type': 'required_field',
                    'message': 'Data Type is required',
                    'sheet': 'Fields',
                    'row': row_idx,
                    'column': headers.get('data_type')
                })
            elif data_type not in VALID_DATA_TYPES:
                field_errors.append({
                    'type': 'invalid_value',
                    'message': f'Invalid data type "{data_type}". Valid types: {", ".join(VALID_DATA_TYPES.keys())}',
                    'sheet': 'Fields',
                    'row': row_idx,
                    'column': headers.get('data_type')
                })
            
            # Generate or validate API name
            api_name = row_data.get('api_name') or self.generate_api_name(field_label)
            api_name = api_name.lower()
            
            # Check for duplicate API names
            if api_name in seen_api_names:
                field_errors.append({
                    'type': 'duplicate',
                    'message': f'Duplicate field API name: {api_name}',
                    'sheet': 'Fields',
                    'row': row_idx,
                    'column': headers.get('api_name')
                })
            seen_api_names.add(api_name)
            
            # Validate picklist values
            if data_type == 'picklist' or data_type == 'multipicklist':
                picklist_values = row_data.get('picklist_values', '')
                if not picklist_values:
                    field_errors.append({
                        'type': 'required_field',
                        'message': 'Picklist Values are required for picklist fields',
                        'sheet': 'Fields',
                        'row': row_idx,
                        'column': headers.get('picklist_values')
                    })
            
            # Validate lookup object
            if data_type == 'lookup':
                lookup_object = row_data.get('lookup_object', '')
                if not lookup_object:
                    field_errors.append({
                        'type': 'required_field',
                        'message': 'Lookup Object is required for lookup fields',
                        'sheet': 'Fields',
                        'row': row_idx,
                        'column': headers.get('lookup_object')
                    })
            
            errors.extend(field_errors)
            
            # Build field data
            required = str(row_data.get('required', '')).lower() in ('true', 'yes', '1', 'x')
            
            field_config = {
                'name': api_name,
                'label': field_label,
                'type': data_type,
                'required': required,
                'default_value': row_data.get('default_value'),
                'editable': True,
                'system': False
            }
            
            # Add picklist values
            if data_type in ('picklist', 'multipicklist'):
                picklist_str = row_data.get('picklist_values', '')
                if picklist_str:
                    values = [v.strip() for v in str(picklist_str).split(',') if v.strip()]
                    field_config['picklist_values'] = values
            
            # Add lookup object
            if data_type == 'lookup':
                field_config['lookup_object'] = row_data.get('lookup_object', '').lower()
            
            fields_data.append(field_config)
        
        return fields_data, errors
    
    async def validate_object(self, tenant_id: str, object_data: Dict[str, Any], fields_data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Validate object data against existing objects in the database.
        """
        errors = []
        
        # Check if object API name already exists
        existing = await self.objects_collection.find_one({
            'tenant_id': tenant_id,
            'object_name': object_data.get('object_name', '').lower()
        })
        
        if existing:
            errors.append({
                'type': 'duplicate_object',
                'message': f'Object with API name "{object_data.get("object_name")}" already exists',
                'sheet': 'Object',
                'row': 2,
                'column': None
            })
        
        # Validate lookup references
        for idx, field in enumerate(fields_data):
            if field.get('type') == 'lookup' and field.get('lookup_object'):
                lookup_obj = field['lookup_object']
                # Check if lookup object exists (skip user which is always valid)
                if lookup_obj != 'user':
                    lookup_exists = await self.objects_collection.find_one({
                        'tenant_id': tenant_id,
                        'object_name': lookup_obj
                    })
                    if not lookup_exists:
                        errors.append({
                            'type': 'invalid_lookup',
                            'message': f'Lookup object "{lookup_obj}" does not exist',
                            'sheet': 'Fields',
                            'row': idx + 2,
                            'column': None
                        })
        
        return errors
    
    async def create_object_from_excel(
        self,
        tenant_id: str,
        user_id: str,
        object_data: Dict[str, Any],
        fields_data: List[Dict[str, Any]]
    ) -> Dict[str, Any]:
        """
        Create the custom object and its fields from parsed Excel data.
        This is transactional - if any part fails, everything is rolled back.
        """
        now = datetime.now(timezone.utc).isoformat()
        
        # Build fields dictionary
        fields = {}
        
        # Add user-defined fields
        for field in fields_data:
            field_name = field['name']
            fields[field_name] = {
                'label': field['label'],
                'type': field['type'],
                'required': field.get('required', False),
                'editable': True,
                'system': False
            }
            
            if field.get('default_value'):
                fields[field_name]['default_value'] = field['default_value']
            
            if field.get('picklist_values'):
                fields[field_name]['picklist_values'] = field['picklist_values']
            
            if field.get('lookup_object'):
                fields[field_name]['lookup_object'] = field['lookup_object']
        
        # Add system fields
        for field_name, field_config in SYSTEM_FIELDS.items():
            if field_name not in fields:
                fields[field_name] = field_config.copy()
        
        # Add name field if not present (primary display field)
        if 'name' not in fields:
            fields['name'] = {
                'label': 'Name',
                'type': 'text',
                'required': True,
                'editable': True,
                'system': False
            }
        
        # Build the object document
        object_doc = {
            'id': str(uuid.uuid4()),
            'tenant_id': tenant_id,
            'object_name': object_data['object_name'].lower(),
            'object_label': object_data['object_label'],
            'object_plural': object_data.get('object_plural', object_data['object_label'] + 's'),
            'description': object_data.get('description', ''),
            'icon': object_data.get('icon', 'file'),
            'is_custom': True,
            'is_active': True,
            'fields': fields,
            'created_at': now,
            'created_by': user_id,
            'updated_at': now,
            'updated_by': user_id,
            'import_source': 'excel'
        }
        
        try:
            # Insert the object
            await self.objects_collection.insert_one(object_doc)
            
            # Create default layouts for the new object
            await self._create_default_layouts(tenant_id, user_id, object_doc)
            
            logger.info(f"Created custom object '{object_data['object_name']}' from Excel for tenant {tenant_id}")
            
            return {
                'success': True,
                'object': {
                    'id': object_doc['id'],
                    'object_name': object_doc['object_name'],
                    'object_label': object_doc['object_label'],
                    'object_plural': object_doc['object_plural'],
                    'field_count': len(fields),
                    'fields': list(fields.keys())
                }
            }
            
        except Exception as e:
            logger.error(f"Failed to create object from Excel: {str(e)}")
            # Attempt rollback
            await self.objects_collection.delete_one({
                'tenant_id': tenant_id,
                'object_name': object_data['object_name'].lower()
            })
            raise e
    
    async def _create_default_layouts(self, tenant_id: str, user_id: str, object_doc: Dict[str, Any]):
        """Create default New and Detail layouts for the custom object"""
        now = datetime.now(timezone.utc).isoformat()
        object_name = object_doc['object_name']
        object_label = object_doc['object_label']
        fields = object_doc['fields']
        
        # Get user-editable fields for layouts (exclude system fields)
        layout_fields = [
            field_name for field_name, config in fields.items()
            if not config.get('system', False) and field_name not in ['created_at', 'created_by', 'updated_at', 'updated_by']
        ]
        
        # New Layout
        new_layout = {
            'id': str(uuid.uuid4()),
            'tenant_id': tenant_id,
            'object_name': object_name,
            'layout_name': f'{object_label} New Layout',
            'api_name': f'{object_name}_New_Layout',
            'description': f'Default new record layout for {object_label}',
            'page_type': 'new',
            'is_system': True,
            'is_active': True,
            'template_type': 'form',
            'sections': [
                {
                    'name': f'{object_label} Information',
                    'columns': 2,
                    'fields': layout_fields[:12]  # First 12 fields
                }
            ],
            'required_fields': [f for f, c in fields.items() if c.get('required', False)],
            'default_values': {},
            'created_at': now,
            'created_by': user_id,
            'updated_at': now
        }
        
        # Detail Layout
        # Split fields into sections
        main_fields = layout_fields[:6]
        additional_fields = layout_fields[6:12] if len(layout_fields) > 6 else []
        system_fields = ['created_at', 'created_by', 'updated_at', 'updated_by', 'owner_id']
        
        detail_layout = {
            'id': str(uuid.uuid4()),
            'tenant_id': tenant_id,
            'object_name': object_name,
            'layout_name': f'{object_label} Detail Layout',
            'api_name': f'{object_name}_Detail_Layout',
            'description': f'Default detail layout for {object_label}',
            'page_type': 'detail',
            'is_system': True,
            'is_active': True,
            'template_type': 'two_column',
            'header_fields': main_fields[:4],
            'highlight_fields': [],
            'show_stage_path': False,
            'sections': [
                {
                    'name': f'{object_label} Information',
                    'columns': 2,
                    'fields': main_fields
                }
            ],
            'regions': [
                {
                    'id': 'center',
                    'name': 'Details',
                    'width': 'flex-1',
                    'order': 0,
                    'components': [
                        {
                            'type': 'tabs',
                            'config': {
                                'tabs': [
                                    {
                                        'id': 'details',
                                        'label': 'Details',
                                        'sections': [
                                            {
                                                'name': f'{object_label} Information',
                                                'columns': 2,
                                                'fields': main_fields
                                            },
                                            {
                                                'name': 'Additional Details',
                                                'columns': 2,
                                                'fields': additional_fields
                                            } if additional_fields else None,
                                            {
                                                'name': 'System Information',
                                                'columns': 2,
                                                'fields': [f for f in system_fields if f in fields]
                                            }
                                        ]
                                    }
                                ]
                            }
                        }
                    ]
                },
                {
                    'id': 'right',
                    'name': 'Activity',
                    'width': 'w-80',
                    'order': 1,
                    'components': [
                        {'type': 'activity_timeline', 'config': {'show_tasks': True, 'show_events': True, 'show_emails': True}}
                    ]
                }
            ],
            'created_at': now,
            'created_by': user_id,
            'updated_at': now
        }
        
        # Clean up None values from sections
        for tab in detail_layout['regions'][0]['components'][0]['config']['tabs']:
            tab['sections'] = [s for s in tab['sections'] if s is not None]
        
        # Insert layouts
        layouts_collection = self.db.lightning_page_layouts
        await layouts_collection.insert_many([new_layout, detail_layout])
    
    def generate_sample_template(self) -> bytes:
        """Generate a sample Excel template for object import"""
        output = io.BytesIO()
        workbook = Workbook()
        
        # Object sheet
        obj_sheet = workbook.active
        obj_sheet.title = 'Object'
        obj_sheet.append(['Object Label', 'Plural Label', 'API Name', 'Description', 'Icon'])
        obj_sheet.append(['Invoice', 'Invoices', 'invoice', 'Track customer invoices', 'file-text'])
        
        # Set column widths
        obj_sheet.column_dimensions['A'].width = 15
        obj_sheet.column_dimensions['B'].width = 15
        obj_sheet.column_dimensions['C'].width = 15
        obj_sheet.column_dimensions['D'].width = 30
        obj_sheet.column_dimensions['E'].width = 15
        
        # Fields sheet
        fields_sheet = workbook.create_sheet('Fields')
        fields_sheet.append([
            'Field Label', 'API Name', 'Data Type', 'Required', 
            'Default Value', 'Picklist Values', 'Lookup Object'
        ])
        
        # Sample fields
        sample_fields = [
            ['Invoice Number', 'invoice_number', 'text', 'TRUE', '', '', ''],
            ['Amount', 'amount', 'currency', 'TRUE', '', '', ''],
            ['Invoice Date', 'invoice_date', 'date', 'TRUE', '', '', ''],
            ['Due Date', 'due_date', 'date', 'FALSE', '', '', ''],
            ['Status', 'status', 'picklist', 'TRUE', 'Draft', 'Draft,Sent,Paid,Overdue,Cancelled', ''],
            ['Account', 'account_id', 'lookup', 'FALSE', '', '', 'account'],
            ['Contact', 'contact_id', 'lookup', 'FALSE', '', '', 'contact'],
            ['Description', 'description', 'textarea', 'FALSE', '', '', ''],
            ['Is Paid', 'is_paid', 'boolean', 'FALSE', 'FALSE', '', ''],
        ]
        
        for row in sample_fields:
            fields_sheet.append(row)
        
        # Set column widths
        fields_sheet.column_dimensions['A'].width = 20
        fields_sheet.column_dimensions['B'].width = 20
        fields_sheet.column_dimensions['C'].width = 15
        fields_sheet.column_dimensions['D'].width = 10
        fields_sheet.column_dimensions['E'].width = 15
        fields_sheet.column_dimensions['F'].width = 40
        fields_sheet.column_dimensions['G'].width = 15
        
        workbook.save(output)
        output.seek(0)
        return output.read()
